import json
import requests
import time
import os
import shutil
from snowflake.sqlalchemy import URL
from sqlalchemy import create_engine, text
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv
from openpyxl.styles import Font

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_DIR, INPUT_DIR, OUTPUT_DIR = [os.path.join(BASE_DIR, d) for d in ['config', 'input', 'output']]
MASK_CONFIG_PATH = os.path.join(CONFIG_DIR, 'mask_configuration.json')
CREDENTIALS_PATH = os.path.join(CONFIG_DIR, 'credentials.json')
TABLE_LIST_PATH = os.path.join(INPUT_DIR, 'input_list.txt')
MASK_EXCEL_PATH = os.path.join(OUTPUT_DIR, 'mask_report.xlsx')
MASK_TRACKING_PATH = os.path.join(BASE_DIR, 'tracking_ids.txt')

MAX_COLUMNS_PER_API_CALL = 50

def load_env_config():
    load_dotenv()
    return int(os.getenv("NUM_ROWS", 100)), int(os.getenv("TABLE_CHUNK_SIZE", 10))

def load_credentials(file_path=CREDENTIALS_PATH):
    if not os.path.exists(file_path):
        raise Exception(f"Credentials file not found: {file_path}")
    with open(file_path, 'r') as file:
        return json.load(file)

def load_table_list(file_path):
    if not os.path.exists(file_path):
        return None
    with open(file_path, 'r') as file:
        return [line.strip() for line in file.readlines() if line.strip()]

def get_snowflake_connection(credentials):
    return create_engine(
        URL(
            account=credentials['account'],
            user=credentials['user'],
            password=credentials['password'],
            warehouse=credentials['warehouse'],
            role=credentials['role'],
            application="protecto"
        ),
        connect_args={'client_session_keep_alive': True}
    )

def fetch_data_from_snowflake(engine, table_name, limit, offset):
    query = f"SELECT * FROM {table_name} LIMIT {limit} OFFSET {offset}"
    with engine.connect() as connection:
        result = connection.execute(text(query))
        return list(result.keys()), result.fetchall()

def create_mask_payload(columns, rows, start_row, COLUMN_MAPPING, column_start=0, column_end=None):
    payload = {"mask": []}
    if not columns or not rows:
        return payload
    
    column_end = len(columns) if column_end is None else column_end
    
    for row_number, row in enumerate(rows, start=start_row):
        for position in range(column_start, min(column_end, len(row))):
            value = row[position]
            if value is None:
                continue
                
            column_name = columns[position] if position < len(columns) else f"column_{position}"
            column_info = COLUMN_MAPPING.get(position, {"format": None, "token_name": None})
            entry = {
                "value": str(value),
                "attribute": {
                    "row": row_number,
                    "column": column_name,
                    "column_position": position
                }
            }
            if column_info.get("format"):
                entry["format"] = column_info["format"]
            if column_info.get("token_name"):
                entry["token_name"] = column_info["token_name"]
                
            payload["mask"].append(entry)
    
    return payload

def create_or_append_output_excel(output_path, masked_data, table_name):
    if not table_name:
        return
    
    print(f"Creating output for table: {table_name}")
    table_excel_path = os.path.join(os.path.dirname(output_path), f"{table_name}.xlsx")
    os.makedirs(os.path.dirname(table_excel_path), exist_ok=True)
    
    wb = load_workbook(table_excel_path) if os.path.exists(table_excel_path) else Workbook()
    if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(wb['Sheet'])
    
    sheet_name = "Report"
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(title=sheet_name)
    
    if not masked_data:
        wb.save(table_excel_path)
        return
        
    valid_data = [item for item in masked_data if 
                'attribute' in item and 
                'row' in item['attribute'] and 
                'column_position' in item['attribute']]
                
    if not valid_data:
        wb.save(table_excel_path)
        return
        
    sorted_results = sorted(valid_data, key=lambda x: (x['attribute']['row'], x['attribute']['column_position']))
    existing_headers = {col-1: ws.cell(row=1, column=col).value 
                      for col in range(1, ws.max_column + 1) 
                      if ws.cell(row=1, column=col).value} if sheet_name in wb.sheetnames else {}
    
    max_col_index = max(
        max([res['attribute']['column_position'] for res in sorted_results] or [0]),
        max(existing_headers.keys() or [0])
    )
    
    column_names = [''] * (max_col_index + 1)
    for pos, name in existing_headers.items():
        if 0 <= pos < len(column_names):
            column_names[pos] = name
    
    for res in sorted_results:
        col_idx = res['attribute']['column_position']
        if col_idx < len(column_names) and not column_names[col_idx]:
            column_names[col_idx] = res['attribute'].get('column', f"Column_{col_idx}")

    header_font = Font(bold=True)
    for idx, name in enumerate(column_names):
        if name:
            cell = ws.cell(row=1, column=idx + 1)
            if cell.value is None:
                cell.value = name
            cell.font = header_font

    for result in sorted_results:
        if 'masked_value' in result:
            ws.cell(
                row=result['attribute']['row'] + 2, 
                column=result['attribute']['column_position'] + 1
            ).value = result['masked_value']

    wb.save(table_excel_path)
    print(f"Saved data for table {table_name}")

def process_tracking_ids(base_url, auth_key, output_excel_path, current_table, tracking_file):
    if not os.path.exists(tracking_file):
        return True
        
    with open(tracking_file, 'r') as f:
        tracking_ids = [line.strip() for line in f.readlines() if line.strip()]
        
    if not tracking_ids:
        return True
        
    for tracking_id in tracking_ids:
        try:
            masked_data = track_status(base_url, auth_key, tracking_id)
            if not masked_data:
                print(f"Error: No masked data returned for tracking ID {tracking_id}")
                print(f"Skipping masking for table '{current_table}' due to tracking ID failures. Moving on to the next table.")
                return False
            create_or_append_output_excel(output_excel_path, masked_data, current_table)
            print(f"Masked data stored in excel for {current_table}")
        except Exception as e:
            print(f"Error processing tracking ID {tracking_id}: {str(e)}")
            print(f"Skipping masking for table '{current_table}' due to tracking ID failures. Moving on to the next table.")
            return False
    return True

def load_column_mapping(table_name, config_path=MASK_CONFIG_PATH):
    if not os.path.exists(config_path):
        raise Exception(f"mask_configuration file not found: {config_path}")
    try:
        with open(config_path, 'r') as f:
            file_content = f.read().strip()
            if not file_content:
                print(f"No mask configuration found for '{table_name}'. Proceeding with auto-detection for {table_name}")
                return {}
            all_mappings = json.loads(file_content)
    except json.JSONDecodeError:
        raise Exception(f"Invalid JSON in mask_configuration file.")

    case_insensitive_map = {k.upper(): k for k in all_mappings.keys()}
    table_key_upper = table_name.strip().upper()
    
    if table_key_upper not in case_insensitive_map:
        return {}  

    original_key = case_insensitive_map[table_key_upper]
    mapping = all_mappings[original_key]
    
    return {int(k): {"format": v.get("format"), "token_name": v.get("token_name")} 
            for k, v in mapping.items()}

def call_mask_api(base_url, auth_key, mask_payload):
    headers = {
        'Authorization': f'Bearer {auth_key}',
        'Content-Type': 'application/json',
    }
    response = requests.put(
        f"{base_url}/mask/async", 
        headers=headers, 
        json=mask_payload
    ) 
    response_data = response.json()
    if not response_data.get('success'):
        raise Exception("Error in async mask api: " + str(response_data.get('error', {}).get('message')))
    return response_data

def check_status(base_url, auth_key, tracking_id):
    headers = {
        'Authorization': f'Bearer {auth_key}',
        'Content-Type': 'application/json',
    }
    payload = {"status": [{"tracking_id": tracking_id.strip()}]}
    response = requests.put(
        f"{base_url}/async-status", 
        headers=headers, 
        json=payload
    )
    response_data = response.json()
    if not response_data.get('success'):
        raise Exception("Error checking status: " + str(response_data.get('error', {}).get('message')))
    return response_data
    
def track_status(base_url, auth_key, tracking_id):
    if not tracking_id or not tracking_id.strip():
        return []
    while True:
        status_response = check_status(base_url, auth_key, tracking_id)
        
        if not status_response or 'data' not in status_response or not status_response['data']:
            raise Exception(f"Invalid status response for tracking ID {tracking_id}")
            
        status = status_response['data'][0]['status']
        
        if status == 'SUCCESS':
            if 'result' not in status_response['data'][0]:
                return []
                
            result = status_response['data'][0]['result']
            if not result:
                return []
                
            return [{'attribute': res['attribute'], 'masked_value': res['token_value']} 
                    for res in result if 'attribute' in res and 'token_value' in res]
                    
        elif status == 'FAILED':
            error_message = status_response['data'][0].get('error', {}).get('message', 'Unknown error')
            raise Exception(f"Processing failed for tracking ID {tracking_id}: {error_message}")
        elif status in ['IN-PROGRESS', 'PENDING']:
            time.sleep(5)
        else:
            raise Exception(f"Unexpected status '{status}' for tracking ID {tracking_id}")
        
def validate_api_response(response, offset):
    if not response:
        return "No response received from API for batch at offset {}".format(offset)
    if not response.get('data') or len(response['data']) == 0:
        return "Invalid or empty 'data' in API response for batch at offset {}".format(offset)
    if 'tracking_id' not in response['data'][0] or not response['data'][0]['tracking_id']:
        return "Missing or empty 'tracking_id' in API response for batch at offset {}".format(offset)
    return None

def process_mask_request(columns, rows, offset, COLUMN_MAPPING, base_url, auth_key, tracking_file, col_start=0, col_end=None):
    payload = create_mask_payload(columns, rows, offset, COLUMN_MAPPING, col_start, col_end)
    if not payload["mask"]:
        return None
    
    response = call_mask_api(base_url, auth_key, payload)
    error_message = validate_api_response(response, offset)
    if error_message:
        print(error_message)
        return None
    
    tracking_id = response['data'][0]['tracking_id']
    with open(tracking_file, 'a') as f:
        f.write(f"{tracking_id}\n")
    
    return tracking_id

def mask_processing(base_url, auth_key, columns, rows, offset, table, tracking_file):
    try:
        COLUMN_MAPPING = load_column_mapping(table, MASK_CONFIG_PATH)
        if COLUMN_MAPPING and columns:
            max_col_index = len(columns) - 1
            out_of_range = [pos for pos in COLUMN_MAPPING.keys() if pos > max_col_index]
            if out_of_range and offset == 0:  
                raise Exception(f"Table {table} has {len(columns)} columns, but config specified out-of-range indices: {out_of_range}")
        
        tracking_dir = os.path.dirname(tracking_file)
        if tracking_dir and not os.path.exists(tracking_dir):
            os.makedirs(tracking_dir)
        
        if len(columns) > MAX_COLUMNS_PER_API_CALL:
            tracking_ids = []
            for col_start in range(0, len(columns), MAX_COLUMNS_PER_API_CALL):
                col_end = min(col_start + MAX_COLUMNS_PER_API_CALL, len(columns))
                print(f"Masking started for columns {col_start+1} to {col_end} for table {table} at offset {offset}")
                tracking_id = process_mask_request(columns, rows, offset, COLUMN_MAPPING, 
                                                base_url, auth_key, tracking_file, 
                                                col_start, col_end)
                if tracking_id:
                    tracking_ids.append(tracking_id)
            
            return tracking_ids[-1] if tracking_ids else None
        else:
            return process_mask_request(columns, rows, offset, COLUMN_MAPPING, 
                                    base_url, auth_key, tracking_file)
    except Exception as e:
        raise Exception(f"Error processing mask operation for table {table}: {str(e)}")
    
def validate_input_parameters(base_url, num_rows, credentials, table_list_file):
    if not base_url:
        raise Exception("API base URL is required")
        
    if num_rows <= 0:
        raise Exception(f"Invalid number of rows: {num_rows}")
    
    if not credentials:
        raise Exception("Failed to load valid credentials")
    
    required_fields = ['account', 'user', 'password', 'warehouse', 'role', 'protecto_api_key']
    missing_fields = [field for field in required_fields if field not in credentials]
    
    if missing_fields:
        raise Exception(f"Missing required fields in credentials file: {', '.join(missing_fields)}")
        
    auth_key = credentials.get('protecto_api_key')
    if not auth_key:
        raise Exception("Protecto API key missing in credentials file")
    
    table_list = load_table_list(table_list_file)
    if not table_list:
        raise Exception("No tables found in input_list file")
    
    return {
        "credentials": credentials,
        "auth_key": auth_key,
        "table_list": table_list
    }

def main(base_url, table_list_file, num_rows, TABLE_CHUNK_SIZE, output_excel_path, tracking_file):
    try:
        if os.path.exists(OUTPUT_DIR):
            shutil.rmtree(OUTPUT_DIR)

        credentials = load_credentials()
        validation_result = validate_input_parameters(base_url, num_rows, credentials, table_list_file)
         
        credentials = validation_result["credentials"]
        auth_key = validation_result["auth_key"]
        table_list = validation_result["table_list"]
       
        engine = get_snowflake_connection(credentials)
        print(f"Masking started: max {NUM_ROWS} rows fetched, {TABLE_CHUNK_SIZE} row per API call, {MAX_COLUMNS_PER_API_CALL} columns per call")
        
        for table in table_list:
            for path in [tracking_file, output_excel_path]:
                if os.path.exists(path):
                    os.remove(path)
                
            row_count = 0
            for offset in range(0, num_rows, TABLE_CHUNK_SIZE):
                chunk_size = min(TABLE_CHUNK_SIZE, num_rows - offset)
                
                columns, rows = fetch_data_from_snowflake(engine, table, chunk_size, offset)                 
                if not rows:
                    break
                
                row_count += len(rows)
                mask_processing(base_url, auth_key, columns, rows, offset, table, tracking_file)
                print(f"Mask API call completed for table '{table}' at row offset {offset + 1}")

            if row_count > 0:
                if not process_tracking_ids(base_url, auth_key, output_excel_path, table, tracking_file):
                    print(f"Masking failed for table {table} due to tracking ID failures. Proceeding with next table")

        print("All table masking operations completed.")

    except Exception as e:
        raise Exception(f"Error in masking process: {str(e)}")
    
if __name__ == '__main__':
    NUM_ROWS, TABLE_CHUNK_SIZE = load_env_config()
    BASE_URL = "https://qa.protecto.ai/api/vault"
    main(BASE_URL, TABLE_LIST_PATH, NUM_ROWS, TABLE_CHUNK_SIZE, MASK_EXCEL_PATH, MASK_TRACKING_PATH)